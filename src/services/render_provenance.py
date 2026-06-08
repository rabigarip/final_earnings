"""Provenance .xlsx sidecar — one row per number on the deck.

Walks `canonical_store` for the ticker, plus the memo computed values
and the macro snapshot, and emits an Excel workbook the analyst can
open alongside the .pptx to audit every figure: source provider,
URL, the year/quarter the value belongs to, and when it was fetched.

This file is the answer to "where did each number come from, and for
which period?" — the single sidecar the user asked for. LLM-generated
prose cells are marked `LLM (Gemini)` with the upstream numeric anchor
ID list rather than left ambiguous.

Layout:
  Sheet "Provenance"
    Slide | Section | Metric | Value | Source | Source URL | Data Period | Fetched At | Notes
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

from src.services.canonical_store import (
    get_all_fields, get_observations_by_provider,
)


# ── Field → slide / section / human label mapping ────────────────
#
# Maps canonical_store field names to where they appear on the deck.
# Fields not in the map are still surfaced (as "Other / Other") so the
# analyst sees the full inventory; this only controls grouping.
_FIELD_MAP: dict[str, tuple[str, str, str]] = {
    # field name              (slide, section, human label)
    "company_profile":        ("Slide 1", "Header", "Company profile"),
    "quote":                  ("Slide 1", "Key Data", "Quote (price + market cap bundle)"),
    "current_price":          ("Slide 1", "Key Data", "Last close"),
    "market_cap":             ("Slide 1", "Key Data", "Market cap"),
    "dividend_yield":         ("Slide 1", "Key Data", "Dividend yield (TTM)"),
    "target_price":           ("Slide 1", "Analyst Consensus", "Average target price"),
    "rating_split":           ("Slide 1", "Analyst Consensus", "Rating split (buy/hold/sell)"),
    "consensus_target":       ("Slide 1", "Analyst Consensus", "Average target price"),
    "valuation_forward":      ("Slide 1", "Key Data + Slide 2 Table",
                                "Forward P/E + Q+1 forecasts"),
    "fifty_two_week_high":    ("Slide 1", "Key Data", "52-week high"),
    "fifty_two_week_low":     ("Slide 1", "Key Data", "52-week low"),
    "valuation_historical":   ("Slide 3", "P/E Chart", "Historical P/E series"),
    "historical_prices":      ("Slide 1", "52-Week Price + Recent Performance",
                                "Daily closes + 52w range + performance buckets"),
    "income_statement_quarterly": ("Slide 3", "Earnings History Chart",
                                     "Per-quarter actual vs estimate surprise track"),
    "income_statement_annual":    ("Slide 2", "Estimates Table (history)",
                                     "Annual revenue / EBITDA / NI history"),
    "broker_actions":         ("Slide 2", "Catalysts (track-record anchor)", "Recent broker actions"),
    "dividends_payments":     ("Slide 1", "Key Data", "Dividend yield + payment history"),
    "earnings_calendar":      ("Slide 1", "Key Data", "Earnings date"),
}


# ── Helpers ──────────────────────────────────────────────────────

def _value_repr(v: Any) -> str:
    """Compact text rendering of a canonical_value for the Value column.
    Dicts and lists are collapsed to their key list so the row stays
    legible — the analyst can drill into the raw .json fixtures if a
    deeper view is needed."""
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        if abs(v) >= 1e9:
            return f"{v:,.2f}"
        if abs(v) >= 1e6:
            return f"{v:,.0f}"
        if abs(v) >= 1:
            return f"{v:,.2f}"
        return f"{v:.4f}"
    if isinstance(v, dict):
        keys = list(v.keys())[:6]
        more = " …" if len(v) > 6 else ""
        return f"<dict: {', '.join(keys)}>{more}"
    if isinstance(v, list):
        return f"<list[{len(v)}]>"
    s = str(v)
    return s if len(s) <= 80 else s[:77] + "…"


def _source_url(provider: str, ticker: str) -> str:
    """Best-effort URL for the canonical source. We don't keep per-cell
    URLs in canonical_store (the raw_observations table holds those),
    so this gives the analyst the right page to land on for a given
    provider."""
    p = (provider or "").lower()
    if "investing" in p:
        return "https://www.investing.com (per-equity slug, see data/investing/)"
    if "marketscreener" in p:
        return "https://www.marketscreener.com"
    if "yahoo" in p or "yfinance" in p:
        return f"https://finance.yahoo.com/quote/{ticker}"
    if "bloomberg" in p:
        return "Analyst Bloomberg export (data/bloomberg/)"
    if p in ("imf", "imf weo"):
        return "https://www.imf.org/external/datamapper"
    if p in ("wb", "world bank", "worldbank"):
        return "https://api.worldbank.org/v2"
    if "gemini" in p or "llm" in p:
        return "https://aistudio.google.com (Gemini API)"
    return ""


def _data_period_from_value(field: str, value: Any) -> str:
    """Pull a human period tag from the value payload when present.
    Different providers store period tags under different keys —
    this is the de-facto union over what we've seen in the wild."""
    if not isinstance(value, dict):
        return ""
    # Macro fields: explicit year tag with source label.
    if field == "company_profile":
        bits = []
        if value.get("macro_year"):
            bits.append(f"macro {value.get('macro_year')}")
        if value.get("gdp_growth_fcst_year"):
            bits.append(f"IMF GDP {value.get('gdp_growth_fcst_year')}F")
        if value.get("inflation_fcst_year"):
            bits.append(f"IMF infl {value.get('inflation_fcst_year')}F")
        if bits:
            return ", ".join(bits)
    # Forecast bundle: surface fy1_year / next_q_period as the period anchor.
    if field == "valuation_forward":
        bits = []
        if value.get("next_q_period"):
            bits.append(str(value["next_q_period"]))
        if value.get("fy1_year"):
            bits.append(f"FY{value['fy1_year']}E")
        return ", ".join(bits)
    # Generic keys some providers stamp.
    for k in ("period", "as_of", "year", "fiscal_year", "report_year", "asof"):
        if value.get(k):
            return str(value[k])
    return ""


# ── Main entry ───────────────────────────────────────────────────

def _explode_dict_value(slide: str, section: str, label_prefix: str,
                          d: dict, src: str, src_url: str,
                          fetched_at: str, notes: str) -> list[list[str]]:
    """Expand a dict-valued canonical_store cell into one row per scalar
    sub-key, so the analyst sees every number on its own line instead of
    a `<dict: ...>` placeholder. Skips nested dicts/lists (the parent row
    still emits as a summary)."""
    out_rows = []
    SKIP_KEYS = {"summary"}   # long-form prose, not a numeric provenance row
    for k, v in d.items():
        if k in SKIP_KEYS:
            continue
        if v is None:
            continue
        if isinstance(v, (dict, list)):
            # Skip nested structures from the explosion; the parent row
            # surfaces them with a `<dict>` / `<list>` placeholder so the
            # analyst knows to inspect them in raw fixtures.
            continue
        # Format numbers compactly; leave strings as-is.
        if isinstance(v, (int, float)):
            disp = (f"{v:,.4f}" if abs(v) < 1 else
                    f"{v:,.2f}" if abs(v) < 1e6 else
                    f"{v:,.0f}")
        else:
            s = str(v)
            disp = s if len(s) <= 80 else s[:77] + "…"
        out_rows.append([
            slide, section, f"{label_prefix} · {k}",
            disp, src, src_url, "", fetched_at, notes,
        ])
    return out_rows


def write_provenance_xlsx(ticker: str, out_path: Path,
                            memo_data: Optional[dict] = None,
                            payload: Any = None,
                            peer_rows: Optional[list[dict]] = None,
                            historical_override: Optional[dict] = None,
                            ) -> Optional[Path]:
    """Walk canonical_store + memo + macro for `ticker` and write an
    .xlsx audit sheet to `out_path`. Returns the path on success, None
    if openpyxl is unavailable (kept optional so the deck still ships).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Provenance"
    headers = ["Slide", "Section", "Metric", "Value", "Source",
                 "Source URL", "Data Period", "Fetched At", "Notes"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0D1117")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    rows: list[list[str]] = []

    # 0. Deck-readiness scorecard — self-reports how grounded this deck is
    #    (data-completeness tier + what's missing) so the analyst sees at a
    #    glance whether the commentary is BKMB-grade or thin, and what to
    #    curate. Best-effort; never blocks the sidecar.
    try:
        from src.services.deck_scorecard import score_ticker as _score
        _sc = _score(ticker)
        rows.append([
            "All Slides", "Deck Readiness",
            f"Data-completeness tier ({_sc['score']}/100)",
            _sc["tier"], "Derived", "src/services/deck_scorecard.py",
            "", "",
            ("Nothing missing — deck-ready" if not _sc["missing"]
             else "To improve: " + "; ".join(_sc["missing"][:4])),
        ])
    except Exception:
        pass

    # 0b. Grounding currency — is the FY baseline the latest the company has
    #     filed, computed against its reporting calendar for TODAY? Makes the
    #     baseline's vintage an explicit, honest fact on the audit trail.
    try:
        from src.services.reporting_calendar import grounding_staleness as _gs
        _cur = _gs(ticker)
        if _cur.get("have_grounded_fy"):
            rows.append([
                "All Slides", "Deck Readiness", "Grounding currency",
                ("CURRENT" if _cur["annual_current"] else "STALE"),
                "Derived", "src/services/reporting_calendar.py",
                _cur.get("disclosed_annual") or "", "",
                _cur.get("status") or "",
            ])
    except Exception:
        pass

    # 1. Every canonical_store cell for the ticker — parent summary row
    #    plus per-sub-key explosion for dict-valued cells so every number
    #    on the deck is individually traceable.
    cv = get_all_fields(ticker)
    for field, c in sorted(cv.items()):
        slide, section, label = _FIELD_MAP.get(field, ("Other", "Other", field))
        period = _data_period_from_value(field, c.value)
        src = (c.canonical_source or "").strip()
        src_url = _source_url(c.canonical_source, ticker)
        fetched_at = c.last_refreshed_at.strftime("%Y-%m-%d %H:%M UTC")
        notes = (c.notes or "").strip()
        rows.append([
            slide, section, label,
            _value_repr(c.value), src, src_url,
            period, fetched_at, notes,
        ])
        # Dict explosion — one row per scalar sub-key. Keeps the analyst
        # from having to open raw JSON to see e.g. `current_price`,
        # `market_cap`, `mean_target_price`, `fwd_pe`, individually.
        if isinstance(c.value, dict):
            rows.extend(_explode_dict_value(slide, section, label,
                                              c.value, src, src_url,
                                              fetched_at, notes))
        # Historical price series: surface 52w stats (high / low / last /
        # period return) since the slide shows them and the full daily
        # series is too noisy to print row-by-row.
        if field == "historical_prices" and isinstance(c.value, list) and c.value:
            try:
                vals = [float(pt["close"]) for pt in c.value
                        if isinstance(pt, dict) and isinstance(pt.get("close"), (int, float))]
                if len(vals) >= 2:
                    hi = max(vals); lo = min(vals)
                    first = vals[0]; last = vals[-1]
                    ret_pct = (last - first) / first * 100.0 if first else None
                    stats = [
                        ("52w high",  f"{hi:,.4f}" if hi < 1 else f"{hi:,.2f}"),
                        ("52w low",   f"{lo:,.4f}" if lo < 1 else f"{lo:,.2f}"),
                        ("Last close", f"{last:,.4f}" if last < 1 else f"{last:,.2f}"),
                        ("# observations", str(len(vals))),
                    ]
                    if ret_pct is not None:
                        stats.append(("52w return %", f"{ret_pct:+.2f}%"))
                    for k, v in stats:
                        rows.append([slide, section, f"{label} · {k}",
                                       v, src, src_url, "", fetched_at, ""])
            except (KeyError, ValueError, TypeError):
                pass
        # Quarterly surprise track: one row per (period, eps_surprise_pct)
        if (field == "income_statement_quarterly" and isinstance(c.value, dict)
            and isinstance(c.value.get("surprise_history"), list)):
            for r in (c.value["surprise_history"] or [])[:8]:
                if not isinstance(r, dict): continue
                p = r.get("period") or r.get("date") or ""
                sp = r.get("eps_surprise_pct")
                ra = r.get("revenue_actual")
                re_ = r.get("revenue_estimate")
                if isinstance(sp, (int, float)) and p:
                    rows.append([slide, section, f"{label} · {p} EPS surprise",
                                   f"{sp:+.2f}%", src, src_url, str(p),
                                   fetched_at, ""])
                if isinstance(ra, (int, float)) and isinstance(re_, (int, float)) and p:
                    rows.append([slide, section, f"{label} · {p} Revenue act/est",
                                   f"{ra:,.0f} / {re_:,.0f}", src, src_url, str(p),
                                   fetched_at, ""])

    # 1c. Slide-1 derived & live-refreshed fields. The deck shows
    #     numbers that DON'T live in canonical_store as scalars
    #     (UPSIDE TO TARGET, the 6 Recent Performance buckets, etc.).
    #     The analyst opening the .xlsx expects to see every visible
    #     number traced. This section makes that contract complete.
    cv_current_price = (cv.get("current_price").value if cv.get("current_price") else None)
    cv_target = (cv.get("target_price").value if cv.get("target_price") else None)
    cv_target_mean = (cv_target.get("mean") if isinstance(cv_target, dict) else None)
    cv_hist = (cv.get("historical_prices").value if cv.get("historical_prices") else None)
    # The historical price series often arrives via historical_override
    # (Investing.com path for yfinance-blocked GCC names) and is NEVER
    # persisted to canonical_store. Without this fallback the provenance
    # silently omitted the 52-week range and all six Recent Performance
    # buckets — the analyst saw them on the deck but couldn't trace them.
    if not (isinstance(cv_hist, dict) and cv_hist.get("close_series")) \
       and isinstance(historical_override, dict) and historical_override.get("close_series"):
        cv_hist = historical_override

    # UPSIDE TO TARGET — derived computation, no canonical row.
    if (isinstance(cv_current_price, (int, float)) and cv_current_price > 0
        and isinstance(cv_target_mean, (int, float))):
        upside = (cv_target_mean - cv_current_price) / cv_current_price * 100.0
        rows.append([
            "Slide 1", "Analyst Consensus", "Upside to target",
            f"{upside:+.2f}%", "Derived",
            f"(target_mean {cv_target_mean:.4f} − last_close {cv_current_price:.4f}) "
            f"/ last_close × 100",
            "", "", "Computed at render time",
        ])

    # 52-WEEK HIGH / LOW — extracted from historical_prices daily series,
    # not stored as separate scalars. Emit rows for the actual numbers
    # rendered on slide 1's range bar.
    if isinstance(cv_hist, dict) and isinstance(cv_hist.get("close_series"), list):
        vals = [float(pt["close"]) for pt in cv_hist["close_series"]
                if isinstance(pt, dict) and isinstance(pt.get("close"), (int, float))]
        if vals:
            hi = max(vals); lo = min(vals)
            # Source attribution: when the values came from the Investing
            # override (because canonical historical_prices was absent OR
            # held an unusable proxy like the iShares EEM series), attribute
            # to Investing — NOT the stale canonical cell. Otherwise the
            # perf/52wk rows would falsely cite "ishares" for BKMB prices.
            _hp_cell = cv.get("historical_prices")
            _using_override = (cv_hist is historical_override)
            hist_src = ("investing" if _using_override
                        else (_hp_cell.canonical_source if _hp_cell else "investing"))
            hist_src_url = _source_url(hist_src, ticker)
            # True as-of = last dated close in the series (NOT read time).
            _last_date = ""
            try:
                _last_date = max(str(pt.get("date") or "")[:10]
                                  for pt in cv_hist["close_series"]
                                  if isinstance(pt, dict) and pt.get("date"))
            except (ValueError, TypeError):
                _last_date = ""
            hist_fetched = ("" if _using_override
                            else (_hp_cell.last_refreshed_at.strftime("%Y-%m-%d %H:%M UTC")
                                  if _hp_cell else ""))
            _series_note = (f"Derived from close series; latest close {_last_date}"
                             if _last_date else "Derived from close series")
            rows.append(["Slide 1", "Key Data", "52-week high",
                          f"{hi:,.4f}" if hi < 1 else f"{hi:,.2f}",
                          hist_src, hist_src_url, _last_date, hist_fetched,
                          _series_note])
            rows.append(["Slide 1", "Key Data", "52-week low",
                          f"{lo:,.4f}" if lo < 1 else f"{lo:,.2f}",
                          hist_src, hist_src_url, _last_date, hist_fetched,
                          _series_note])
            # Recent Performance buckets (1D/1W/1M/3M/6M/YTD). Derived
            # from the same close series — emit one row each.
            from datetime import datetime as _dtP, timezone as _tzP
            from datetime import timedelta as _tdP
            try:
                pts = [(_dtP.strptime(pt["date"], "%Y-%m-%d").date(), float(pt["close"]))
                        for pt in cv_hist["close_series"]
                        if pt.get("date") and isinstance(pt.get("close"), (int, float))]
                pts.sort()
                if pts:
                    last_d, last_v = pts[-1]
                    def _ret_pct(days: int) -> tuple[str, float] | None:
                        target_d = last_d - _tdP(days=days)
                        # Find the closest <= target_d.
                        prior = [(d, v) for d, v in pts if d <= target_d]
                        if not prior: return None
                        anchor = prior[-1]
                        if anchor[1] == 0: return None
                        return (anchor[0].isoformat(),
                                (last_v - anchor[1]) / anchor[1] * 100.0)
                    def _ytd_ret() -> tuple[str, float] | None:
                        jan1 = _dtP(last_d.year, 1, 1).date()
                        prior = [(d, v) for d, v in pts if d >= jan1]
                        if not prior or prior[0][1] == 0: return None
                        return (prior[0][0].isoformat(),
                                (last_v - prior[0][1]) / prior[0][1] * 100.0)
                    # Prefer Investing's OWN published % (what the deck now
                    # shows and what the analyst sees on the site); fall back
                    # to the close-series computation when a published field
                    # is absent. Keeps provenance == deck == Investing.
                    _pub_perf = cv_hist if isinstance(cv_hist, dict) else {}
                    _published = (str(_pub_perf.get("perf_source") or "")
                                  == "investing_published")
                    buckets = [
                        ("1 Day", "perf_1d", _ret_pct(1)),
                        ("1 Week", "perf_1w", _ret_pct(7)),
                        ("1 Month", "perf_1m", _ret_pct(30)),
                        ("3 Months", "perf_3m", _ret_pct(91)),
                        ("6 Months", "perf_6m", _ret_pct(182)),
                        ("YTD", "perf_ytd", _ytd_ret()),
                    ]
                    for label, pkey, result in buckets:
                        pub_v = _pub_perf.get(pkey)
                        if isinstance(pub_v, (int, float)):
                            note = ("Investing.com published figure"
                                    + (f" (updated {str(_pub_perf.get('perf_updated_at'))[:10]})"
                                       if _pub_perf.get("perf_updated_at") else ""))
                            rows.append([
                                "Slide 1", "Recent Performance", label,
                                f"{float(pub_v):+.2f}%",
                                hist_src, hist_src_url, "", hist_fetched, note,
                            ])
                        elif result is not None:
                            anchor_date, ret = result
                            rows.append([
                                "Slide 1", "Recent Performance", label,
                                f"{ret:+.2f}%",
                                hist_src, hist_src_url, "", hist_fetched,
                                f"vs close on {anchor_date}",
                            ])
            except (ValueError, TypeError):
                pass

    # EARNINGS DATE — pulled from memo_data or company / calendar
    # observations; surface explicitly as a slide-1 cell.
    earnings_date_val = None
    earnings_date_src = ""
    if memo_data and memo_data.get("next_q_report_date"):
        earnings_date_val = memo_data["next_q_report_date"]
        earnings_date_src = (memo_data.get("next_q_report_date_source")
                              or "earnings_calendar")
    elif cv.get("earnings_calendar") and isinstance(cv["earnings_calendar"].value, dict):
        ec_val = cv["earnings_calendar"].value
        earnings_date_val = ec_val.get("next_earnings_date") or ec_val.get("date")
        earnings_date_src = cv["earnings_calendar"].canonical_source
    if earnings_date_val:
        rows.append([
            "Slide 1", "Key Data", "Earnings date",
            str(earnings_date_val),
            earnings_date_src or "—",
            _source_url(earnings_date_src, ticker),
            "", "", "Next reporting date per source feed",
        ])

    # P/E — the slide-1 chip shows a FORWARD P/E; the peer table and
    # highlights cite a TRAILING P/E. Both must be traceable, and because
    # sources disagree on P/E (Investing's EPS-implied trailing vs MS's
    # published multiples), the note records the source explicitly.
    import re as _rePE
    _cur_year_pe = datetime.now().year

    def _pe_from_hist(want_forward: bool):
        """Pick a P/E off the MS valuation_historical series — the nearest
        FUTURE fiscal year (forward) or the nearest PAST/current year
        (trailing). Returns (pe, year, source) or (None, None, '')."""
        vh = cv.get("valuation_historical")
        d = vh.value if (vh and isinstance(vh.value, dict)) else {}
        periods = d.get("periods") or []
        pes = d.get("pe") or []
        if len(periods) != len(pes):
            return None, None, ""
        # Forward = nearest FUTURE/current FY (the FY26E multiple).
        # Trailing = most recent COMPLETED FY (strictly past); this is the
        # ~11.1x the peer table / exec summary cite, distinct from forward.
        best = None
        for p, pe in zip(periods, pes):
            m = _rePE.search(r"(\d{4})", str(p) or "")
            if not m or not isinstance(pe, (int, float)) or pe <= 0:
                continue
            yr = int(m.group(1))
            if want_forward:
                if yr < _cur_year_pe:
                    continue
                rank = (yr - _cur_year_pe)           # smallest future first
            else:
                if yr >= _cur_year_pe:
                    continue
                rank = (_cur_year_pe - yr)            # most recent past first
            if best is None or rank < best[2]:
                best = (pe, yr, rank)
        # Trailing fallback: if no strictly-past FY exists, accept current FY.
        if best is None and not want_forward:
            for p, pe in zip(periods, pes):
                m = _rePE.search(r"(\d{4})", str(p) or "")
                if m and isinstance(pe, (int, float)) and pe > 0 and int(m.group(1)) == _cur_year_pe:
                    best = (pe, _cur_year_pe, 0); break
        if best is None:
            return None, None, ""
        return best[0], best[1], (vh.canonical_source if vh else "marketscreener")

    # Forward P/E: valuation_forward.pe_fy1 first, else MS historical series.
    val_fwd = cv.get("valuation_forward")
    fwd = val_fwd.value if (val_fwd and isinstance(val_fwd.value, dict)) else {}
    pe_fy1 = fwd.get("pe_fy1") if isinstance(fwd, dict) else None
    fy1_year = fwd.get("fy1_year") if isinstance(fwd, dict) else None
    if isinstance(pe_fy1, (int, float)) and fy1_year:
        rows.append([
            "Slide 1", "Key Data", f"P/E (FY{str(fy1_year)[-2:]}E)",
            f"{pe_fy1:.2f}x",
            (val_fwd.canonical_source if val_fwd else "—"),
            _source_url(val_fwd.canonical_source if val_fwd else "", ticker),
            f"FY{fy1_year}", "",
            "Forward P/E from valuation_forward bundle (price ÷ FY1 EPS)",
        ])
    else:
        _pe_f, _yr_f, _src_f = _pe_from_hist(want_forward=True)
        if _pe_f is not None:
            rows.append([
                "Slide 1", "Key Data", f"P/E (FY{str(_yr_f)[-2:]}E)",
                f"{_pe_f:.2f}x", _src_f, _source_url(_src_f, ticker),
                f"FY{_yr_f}", "",
                "Forward P/E from MarketScreener valuation series",
            ])

    # Trailing P/E (peer table + highlights anchor): nearest past/current
    # FY off the MS historical series. Documents the 11.1x-class number.
    _pe_t, _yr_t, _src_t = _pe_from_hist(want_forward=False)
    if _pe_t is not None:
        rows.append([
            "Slide 3", "Peer Comparables", "Subject trailing P/E",
            f"{_pe_t:.2f}x", _src_t, _source_url(_src_t, ticker),
            f"FY{_yr_t}", "",
            "Trailing P/E from MarketScreener valuation series",
        ])

    # 1d. LIVE QUOTE RECORD — when the pipeline pre-refreshed volatile
    # fields via yfinance live, surface the freshness anchor so the
    # analyst knows which numbers are intraday-current vs cached.
    lq_rec = (memo_data or {}).get("live_quote_record") if memo_data else None
    if isinstance(lq_rec, dict) and lq_rec.get("ok"):
        fetched = lq_rec.get("fetched_at", "")
        fields_str = ", ".join(lq_rec.get("fields") or [])
        rows.append([
            "Slide 1", "Data Freshness", "Live quote (yfinance)",
            f"Refreshed: {', '.join(lq_rec.get('fields') or []) or '—'}",
            "yfinance-live",
            "https://finance.yahoo.com",
            "", fetched,
            f"Live price={lq_rec.get('price')} · "
            f"mcap={lq_rec.get('market_cap')} · "
            f"52w {lq_rec.get('fifty_two_week_low')}-{lq_rec.get('fifty_two_week_high')}",
        ])

    # 1e. BLOOMBERG OPT-IN MANIFEST — when the analyst has enabled
    # Bloomberg for this ticker, document that explicitly. When
    # NOT enabled, document that too (so the analyst knows the deck
    # ran on free-source data only — no silent override).
    try:
        from pathlib import Path as _PB
        import json as _jsonB
        manifest_path = _PB("data/bloomberg") / f"{ticker}.manifest.json"
        if manifest_path.is_file():
            try:
                m = _jsonB.loads(manifest_path.read_text())
                if m.get("enabled"):
                    rows.append([
                        "All Slides", "Source Overrides",
                        "Bloomberg consensus override",
                        "ENABLED",
                        "Bloomberg upload (analyst-confirmed)",
                        str(manifest_path),
                        "", m.get("uploaded_at", ""),
                        f"Source file: {m.get('source_filename', '?')}",
                    ])
                else:
                    rows.append([
                        "All Slides", "Source Overrides",
                        "Bloomberg consensus override",
                        "DISABLED",
                        "Bloomberg upload present but disabled",
                        str(manifest_path), "", "",
                        "Toggle via /api/v2/bloomberg/manifest/{ticker}/toggle",
                    ])
            except Exception:
                pass
        else:
            rows.append([
                "All Slides", "Source Overrides",
                "Bloomberg consensus override",
                "NOT CONFIGURED",
                "Free-source data only", "", "", "",
                "Analyst can upload a Bloomberg FA xlsx to enable",
            ])
    except Exception:
        pass

    # 2. Memo-derived fields — every Q+1 consensus that fed slide 2 plus
    #    derived YoY / QoQ deltas. Full coverage: revenue, EPS, EBITDA,
    #    Net Income, and the implied upside on slide 1.
    if memo_data:
        nq_src = memo_data.get("next_quarter_consensus_source")
        nq_label = memo_data.get("next_quarter_label") or memo_data.get("next_earnings_label") or ""
        cascade_note = "Cascade: MS quarterly → Investing → Yahoo → MS annual÷4"
        memo_keys_quarterly = [
            ("next_quarter_consensus_revenue", "Q+1 consensus Revenue"),
            ("next_quarter_consensus_eps",     "Q+1 consensus EPS"),
            ("next_quarter_consensus_ebitda",  "Q+1 consensus EBITDA"),
            ("next_quarter_consensus_ni",      "Q+1 consensus Net Income"),
        ]
        for key, human in memo_keys_quarterly:
            v = memo_data.get(key)
            if v is not None:
                rows.append([
                    "Slide 2", "Estimates Table",
                    f"{human} ({nq_label})" if nq_label else human,
                    _value_repr(v),
                    nq_src or "MarketScreener",
                    _source_url(nq_src or "MarketScreener", ticker),
                    nq_label, "", cascade_note,
                ])
        # Implied upside / target reflected in the deck.
        if memo_data.get("implied_upside_pct") is not None:
            rows.append([
                "Slide 1", "Analyst Consensus", "Implied upside %",
                _value_repr(memo_data["implied_upside_pct"]),
                "Derived",
                "(target_mean / quote.price - 1) × 100",
                "", "", "Per-cell formula",
            ])
        # Next-quarter YoY/QoQ derivations (when memo computed them).
        # _compute_memo uses *_growth_pct suffix for revenue/NI; the
        # thesis renderer additionally computes per-metric YoY/QoQ but
        # doesn't write back. Both shapes covered here.
        yoy_qoq_keys = [
            ("next_quarter_yoy_revenue_growth_pct", "Q+1 YoY revenue %"),
            ("next_quarter_yoy_ni_growth_pct",       "Q+1 YoY net income %"),
            ("next_quarter_yoy_revenue",             "Q+1 YoY revenue %"),
            ("next_quarter_yoy_eps",                 "Q+1 YoY EPS %"),
            ("next_quarter_yoy_ebitda",              "Q+1 YoY EBITDA %"),
            ("next_quarter_yoy_ni",                  "Q+1 YoY net income %"),
            ("next_quarter_qoq_revenue",             "Q+1 QoQ revenue %"),
            ("next_quarter_qoq_eps",                 "Q+1 QoQ EPS %"),
            ("next_quarter_qoq_ebitda",              "Q+1 QoQ EBITDA %"),
            ("next_quarter_qoq_ni",                  "Q+1 QoQ net income %"),
        ]
        for key, human in yoy_qoq_keys:
            v = memo_data.get(key)
            if isinstance(v, (int, float)):
                rows.append([
                    "Slide 2", "Estimates Table", human,
                    f"{v:+.2f}%", "Derived",
                    "consensus vs same-Q prior-year (YoY) or prior-Q (QoQ)",
                    nq_label, "", "Computed in _compute_memo / render time",
                ])

    # 3. Peer comparables — emit one row per peer × per shown metric.
    #    Slide 3 shows: COMPANY | TICKER | MCAP | P/E | P/TBV | DIV YIELD |
    #    1Y RETURN. Trace each cell so the analyst can drill into source.
    if peer_rows:
        for pr in peer_rows:
            if not isinstance(pr, dict): continue
            ptick = pr.get("ticker") or pr.get("symbol") or ""
            pname = pr.get("name") or pr.get("company") or ptick
            metrics_to_emit = [
                ("market_cap", "Market cap"),
                ("forward_pe", "Forward P/E"),
                ("price_to_book", "P/TBV"),
                ("price_to_tangible_book", "P/TBV"),
                ("dividend_yield", "Dividend yield %"),
                ("return_1y_pct", "1Y return %"),
            ]
            for src_key, human in metrics_to_emit:
                v = pr.get(src_key)
                if v is None: continue
                rows.append([
                    "Slide 3", "Peer Comparables",
                    f"{pname} ({ptick}) · {human}",
                    _value_repr(v),
                    (pr.get("source") or "yfinance"),
                    _source_url(pr.get("source") or "yfinance", ptick),
                    "", "", "",
                ])

    # 4. Company-disclosed quarterly figures (Phase 1 of the disclosed-
    #    source pipeline). When `data/disclosed/{ticker}.json` exists,
    #    each disclosed quarter is emitted as one row per metric so the
    #    audit trail shows the exact IR-portal PDF the value came from.
    try:
        from src.services.disclosed_loader import load_disclosed
        d_payload = load_disclosed(ticker)
    except Exception:
        d_payload = None
    if d_payload:
        company_name = d_payload.get("company") or ticker
        sd_map = d_payload.get("_source_documents") or {}
        for q in d_payload.get("quarterly", []) or []:
            if not isinstance(q, dict): continue
            period = str(q.get("period") or "").strip()
            if not period: continue
            doc = q.get("source_doc") or ""
            url = sd_map.get(period) or ""
            metric_pairs = [
                ("operating_income", "Operating Income (raw currency)"),
                ("net_interest_income", "Net Interest Income"),
                ("fee_income_net", "Fee & Commission Income (net)"),
                ("net_income", "Net Income"),
                ("eps", "EPS"),
            ]
            for key, human in metric_pairs:
                v = q.get(key)
                if not isinstance(v, (int, float)): continue
                rows.append([
                    "Slide 3", "Earnings History (Disclosed)",
                    f"{human} · {period}",
                    _value_repr(v),
                    f"{company_name} IR",
                    url, period, "",
                    f"Standalone-quarter value from {doc}",
                ])

    # 4b. Disclosed-pipeline coverage assessment. One row that tells
    #     the analyst HOW MUCH of the disclosed track we have for this
    #     ticker — recent quarters captured, freshness, completeness.
    try:
        from src.services.disclosed_status import assess as _disc_assess
        _cov = _disc_assess(ticker)
        if _cov.file_exists:
            rows.append([
                "Slide 3", "Disclosed Coverage",
                f"Coverage summary",
                (f"{_cov.n_quarters} quarters; latest "
                  f"{_cov.most_recent_period} "
                  f"({_cov.days_since_period_end}d old); "
                  f"completeness {_cov.coverage_pct:.0f}%"),
                f"Company IR",
                "(see Earnings History — Disclosed section above)",
                _cov.most_recent_period or "", "",
                ("; ".join(_cov.issues) if _cov.issues
                  else "No issues"),
            ])
    except Exception as _exc:
        pass

    # 4c. Validation findings. Tier-1 (pre-LLM source-data) findings
    #     come from `memo_data["validation_report"]`; tier-2 (post-LLM
    #     numeric trace) comes from `llm["validation_tier2"]`. Both
    #     emit one row per finding so the analyst can audit which
    #     numbers the validator flagged.
    if memo_data and isinstance(memo_data.get("validation_report"), dict):
        for f in memo_data["validation_report"].get("findings", []) or []:
            rows.append([
                "All Slides", "Validation · Tier 1 (pre-LLM)",
                f"{f.get('metric', '?')}",
                (f.get("message") or "")[:120],
                "Derived",
                "src/services/validation_layer.py", "", "",
                f.get("suggested_action") or "",
            ])

    # 4d. MS annual forecasts — emit FY year × metric rows when present.
    #    These back the annual-fallback table on slide 2 and the forward
    #    P/E bars on slide 3.
    annual = None
    if payload is not None:
        msa = getattr(payload, "ms_annual_forecasts", None)
        if isinstance(msa, dict):
            annual = msa.get("annual") if isinstance(msa.get("annual"), dict) else msa
    if annual and isinstance(annual.get("periods"), list):
        periods = annual["periods"]
        for arr_key, human in [("net_sales", "Net sales"),
                                  ("ebitda",    "EBITDA"),
                                  ("net_income", "Net income")]:
            arr = annual.get(arr_key) or []
            for i, p in enumerate(periods):
                if i >= len(arr): continue
                v = arr[i]
                if not isinstance(v, (int, float)): continue
                rows.append([
                    "Slide 2", "Annual Forecasts (MS)",
                    f"{human} · {p}",
                    _value_repr(v),
                    "MarketScreener",
                    "https://www.marketscreener.com",
                    str(p), "",
                    "Annual estimate; per-quarter breakdown not published",
                ])

    # 3. LLM-generated prose (thesis, catalysts, risks, watch_list, highlights).
    # These are not in canonical_store; we surface them explicitly so the
    # analyst sees what the model produced AND that no numerics live here.
    try:
        from src.services.llm_summary import generate_summary
        llm = generate_summary(ticker)
    except Exception:
        llm = None
    if llm:
        as_of = (llm.get("as_of") or "").replace("T", " ")[:19]
        ctx_hash = llm.get("context_hash") or ""
        # Tier-3 QA flags: generic filler / datapoint restatement / banned
        # hedges / length / cross-section repetition. One row per finding so
        # the analyst can see (and the prompt owner can tune) quality issues.
        for q in (llm.get("qa_findings") or []):
            rows.append([
                "Slide 1+2", "QA · Tier 3 (style review)",
                str(q).split(":", 1)[0],
                str(q),
                "Derived", "src/services/llm_summary.py (qa_review)",
                "", as_of, "Non-blocking style flag — review prompt/output",
            ])
        thesis = (llm.get("thesis_paragraph") or "").strip()
        if thesis:
            rows.append([
                "Slide 2", "Investment Thesis", "Executive summary (4-sentence)",
                thesis[:120] + ("…" if len(thesis) > 120 else ""),
                "LLM (Gemini)",
                _source_url("gemini", ticker),
                "", as_of, f"context_hash={ctx_hash}; numeric-trace validated",
            ])
        for i, c in enumerate(llm.get("catalysts") or [], start=1):
            rows.append([
                "Slide 2", "Catalysts", f"Catalyst #{i}",
                str(c)[:120],
                "LLM (Gemini)",
                _source_url("gemini", ticker),
                "", as_of, f"context_hash={ctx_hash}",
            ])
        for i, c in enumerate(llm.get("risks") or [], start=1):
            rows.append([
                "Slide 2", "Risks", f"Risk #{i}",
                str(c)[:120],
                "LLM (Gemini)",
                _source_url("gemini", ticker),
                "", as_of, f"context_hash={ctx_hash}",
            ])
        for i, c in enumerate(llm.get("watch_list") or [], start=1):
            rows.append([
                "Slide 2", "What to Watch", f"Watch #{i}",
                str(c)[:120],
                "LLM (Gemini)",
                _source_url("gemini", ticker),
                "", as_of, f"context_hash={ctx_hash}",
            ])
        for it in (llm.get("highlights") or []):
            cat = (it.get("category") or "").strip().upper()
            body = (it.get("body") or "").strip()
            if cat and body:
                rows.append([
                    "Slide 1", f"Analyst Highlights — {cat}", "Pill body",
                    body[:120],
                    "LLM (Gemini)",
                    _source_url("gemini", ticker),
                    "", as_of, f"context_hash={ctx_hash}",
                ])

        # Tier-2 validation findings (post-LLM numeric-trace).
        v2 = llm.get("validation_tier2") or {}
        for f in (v2.get("findings") or []):
            rows.append([
                "Slide 2", "Validation · Tier 2 (post-LLM)",
                f.get("metric", "?"),
                (f.get("message") or "")[:120],
                "Derived",
                "src/services/validation_layer.py", "", as_of,
                f.get("suggested_action") or "",
            ])

        # Voice eval — score the thesis paragraph against the references.
        try:
            from src.services.voice_eval import evaluate as _voice_eval
            thesis_text = (llm.get("thesis_paragraph") or "").strip()
            if thesis_text:
                vres = _voice_eval(ticker, thesis_text)
                rows.append([
                    "Slide 2", "Voice Eval",
                    "Composite score (vs Apple/JPM/Tesla refs)",
                    f"{vres.composite_score:.3f} ({vres.grade})",
                    "Derived",
                    "src/services/voice_eval.py", "", as_of,
                    f"struct={vres.structural_score:.2f} · "
                    f"lex={vres.lexical_score:.2f} · "
                    f"style={vres.style_score:.2f}"
                    + (" · " + "; ".join(vres.notes[:2]) if vres.notes else ""),
                ])
        except Exception as _exc:
            pass

    # Sort by slide then section for a clean read.
    def _slide_sort(r):
        s = (r[0] or "")
        # "Slide 1" < "Slide 2" < ... < "Other"
        if s.startswith("Slide "):
            try:
                return (0, int(s.split(" ", 1)[1]), r[1])
            except (ValueError, IndexError):
                pass
        return (1, 0, r[1])
    rows.sort(key=_slide_sort)
    for r in rows:
        ws.append(r)

    # Column widths sized to typical content. openpyxl doesn't auto-fit.
    widths = [9, 32, 36, 32, 16, 48, 22, 22, 50]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"

    # Metadata sheet — quick at-a-glance: ticker + run timestamp + counts.
    meta = wb.create_sheet("About")
    meta.append(["Ticker", ticker])
    meta.append(["Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")])
    meta.append(["Rows", len(rows)])
    meta.append([])
    meta.append(["Notes", (
        "Every numeric value on the deck should trace to a row here. "
        "LLM-generated prose has no numeric source — its numbers are "
        "validated against the canonical_store anchors at generation time."
    )])
    for col in ("A", "B"):
        meta.column_dimensions[col].width = 28
    meta["A1"].font = Font(bold=True)
    meta["A2"].font = Font(bold=True)
    meta["A3"].font = Font(bold=True)
    meta["A5"].font = Font(bold=True)
    meta["B5"].alignment = Alignment(wrap_text=True, vertical="top")
    meta.row_dimensions[5].height = 60

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path
